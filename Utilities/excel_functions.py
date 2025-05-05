from openpyxl import load_workbook
class   ExcelFunctions:

    def __init__(self,file_name,sheet_name):
        self.file=file_name
        self.sheet=sheet_name

#Fetch total rows from excel sheet
    def row_count(self):
        workbook=load_workbook(self.file)
        sheet=workbook[self.sheet]
        print("Max row=",sheet.max_row)
        return sheet.max_row

#Fetch total columns from excel sheet
    def column_count(self):
        workbook=load_workbook(self.file)
        sheet=workbook[self.sheet]
        print("Max column=", sheet.max_column)
        return sheet.max_column

#Read data from excel sheet
    def read_data(self,row_number,column_number):
        workbook=load_workbook(self.file)
        sheet=workbook[self.sheet]
        return sheet.cell(row=row_number,column=column_number).value

#Write data to excel sheet
    def write_data(self,row_number,column_number,data):
        workbook=load_workbook(self.file)
        sheet=workbook[self.sheet]
        sheet.cell(row=row_number,column=column_number).value=data
        workbook.save(self.file)
